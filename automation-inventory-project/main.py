import openpyxl

inventory_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inventory_file["Sheet1"]

#List each company with their product count
products_per_supplier = {}
supplier_inventory_count = {}
inventory_less_than_ten = {}

def supplier_product_count():
   for product_row in range(2, product_list.max_row + 1):
      supplier_name = product_list.cell(product_row, 4).value
      if supplier_name in products_per_supplier:
         products_per_supplier[supplier_name] += 1
      else:
         # print("Adding a new Supplier")
         products_per_supplier[supplier_name] = 1
   print(products_per_supplier)

supplier_product_count()

#calculate inventory of each supplier

def total_inventory_value_per_supplier():
   for product_row in range(2, product_list.max_row + 1):
      inventory_value = product_list.cell(product_row, 2).value
      price_value = product_list.cell(product_row, 3).value
      supplier_name = product_list.cell(product_row, 4).value
      if supplier_name in supplier_inventory_count:
         supplier_inventory_count[supplier_name] += inventory_value * price_value
      else:
         supplier_inventory_count[supplier_name] = inventory_value * price_value

   print(supplier_inventory_count)

total_inventory_value_per_supplier()


def products_less_than_ten_inventory():
   for product_row in range(2, product_list.max_row + 1):
      product_number = product_list.cell(product_row, 1).value
      inventory_value = product_list.cell(product_row, 2).value
      if inventory_value < 10:
         inventory_less_than_ten[product_number] = inventory_value

   print(inventory_less_than_ten)

products_less_than_ten_inventory()
